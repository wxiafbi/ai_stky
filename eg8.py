import openpyxl
import datetime
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, Dataset

# 定义数据集类


class ExcelDataset(Dataset):
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet_names = self.workbook.sheetnames

    def __len__(self):
        return len(self.sheet_names)

    def __getitem__(self, index):
        sheet_name = self.sheet_names[index]
        sheet = self.workbook[sheet_name]
        # print(sheet)
        timestamps = []
        values = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            timestamp = datetime.datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
            # print(timestamp)
            value = float(row[1])
            # print(value)
            timestamps.append(timestamp.timestamp())
            values.append(value)
        return timestamps, values, sheet_name

# 定义深度学习模型


class Classifier(nn.Module):
    def __init__(self, input_size, hidden_size, num_classes):
        super(Classifier, self).__init__()
        self.fc1 = nn.Linear(input_size, hidden_size)
        self.relu = nn.ReLU()
        self.fc2 = nn.Linear(hidden_size, num_classes)

    def forward(self, x):
        out = self.fc1(x)
        out = self.relu(out)
        out = self.fc2(out)
        return out


# 加载并预处理数据
dataset = ExcelDataset('训练集.xlsx')
dataloader = DataLoader(dataset, batch_size=1, shuffle=True)
print(dataloader)
# 定义模型参数
input_size = 2
hidden_size = 64
num_classes = len(dataset)

# 初始化模型
model = Classifier(input_size, hidden_size, num_classes)

# 定义训练参数
learning_rate = 0.02
num_epochs = 10

# 定义损失函数和优化器
criterion = nn.CrossEntropyLoss()
optimizer = optim.Adam(model.parameters(), lr=learning_rate)

# 模型训练
for epoch in range(num_epochs):
    total_loss = 0
    for timestamps, values, _ in dataloader:
        # 数据预处理
        timestamps = timestamps[0].float()
        values = values[0].float()
        # print('=============')
        # print(torch.stack((timestamps, values)))
        # print('=============')
        # 前向传播
        output = model(torch.stack((timestamps, values), dim=1))

        # 计算损失
        target = torch.tensor([epoch] * output.size(0), dtype=torch.long)
        loss = criterion(output, target)
        # print(loss)
        # 反向传播和优化
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

        total_loss += loss.item()

    print("Epoch [{}/{}],'|', Loss: {:.4f}".format(epoch +
          1, num_epochs, total_loss))

# 保存训练好的模型权重
torch.save(model.state_dict(), 'model_weights.pth')

# 加载新数据文件并读取单个 sheet
new_file_path = '历史数据.xlsx'
new_workbook = openpyxl.load_workbook(new_file_path)
new_sheet_name = new_workbook.sheetnames[0]
new_sheet = new_workbook[new_sheet_name]

# 读取新数据
new_data = []
for row in new_sheet.iter_rows(min_row=2, values_only=True):
    timestamp = datetime.datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
    value = float(row[1])
    new_data.append((timestamp, value))

# 数据预处理
new_timestamps, new_values = zip(*new_data)
new_timestamps = torch.tensor([timestamp.timestamp()
                              for timestamp in new_timestamps])
new_values = torch.tensor(new_values, dtype=torch.float32)

# 加载已训练好的模型权重
model.load_state_dict(torch.load('model_weights.pth'))
model.eval()

# 模型推理
output = model(torch.stack((new_timestamps, new_values), dim=1))
_, predicted_class = torch.max(output, 1)
predicted_class = predicted_class[0].item()


print("New Data - Predicted Class: {}".format(predicted_class))
