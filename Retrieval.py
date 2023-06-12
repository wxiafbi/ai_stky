import sys

from typing import List

from alibabacloud_iot20180120.client import Client as Iot20180120Client
from alibabacloud_tea_openapi import models as open_api_models
from alibabacloud_iot20180120 import models as iot_20180120_models
from alibabacloud_tea_util import models as util_models
from alibabacloud_tea_util.client import Client as UtilClient

import time_transform
import datetime
import openpyxl

Product_key = ['a1bw1zXB8k4', 'a1Mag4tNOD3',
               'a1gMuOnqrr7', 'a1HyPOJ4feR', 'a1g1PqlfkOo']

Device_name = ['Z065P1', 'Z065P2', 'Z065P6', 'Z065P7']


class Sample:
    '''
    创建客户端
    '''

    def __init__(self):
        pass

    @staticmethod
    def create_client(
        access_key_id: str,
        access_key_secret: str,
    ) -> Iot20180120Client:
        """
        使用AK&SK初始化账号Client
        @param access_key_id:
        @param access_key_secret:
        @return: Client
        @throws Exception
        """
        config = open_api_models.Config(
            # 必填，您的 AccessKey ID,
            access_key_id='LTAI4FbnNz12HT85LUPbrFnU',
            # 必填，您的 AccessKey Secret,
            access_key_secret='UdPN3kNhB8aVJvrrbfIyng5sHupNpA'
        )
        # 访问的域名
        config.endpoint = f'iot.cn-shanghai.aliyuncs.com'
        return Iot20180120Client(config)


class Data_snapshot:
    '''
    读取数据快照
    '''

    def __init__(self, ProductKey: str, DeviceName: str) -> None:
        self.ProductKey = ProductKey
        self.DeviceName = DeviceName

    def PRODUCT(self):
        '''
        设置产品名称
        '''
        return self.ProductKey

    def DEVICE(self):
        '''
        设置设备名称
        '''
        return self.DeviceName

    @staticmethod
    def DE_dict(luck: str):
        DN = []
        DD = {}
        DD['井号'] = luck
        DN.append(DD)
        return DN

    def data_out(self, key_s: str, name_s: str) -> dict:
        DATA_L = []
        DATA_D = {}
        client = Sample.create_client(
            'LTAI4FbnNz12HT85LUPbrFnU', 'UdPN3kNhB8aVJvrrbfIyng5sHupNpA')
        query_device_property_status_request = iot_20180120_models.QueryDevicePropertyStatusRequest(
            product_key=key_s, device_name=name_s)
        runtime = util_models.RuntimeOptions(
            max_attempts=1000, connect_timeout=1000)
        return_value = client.query_device_property_status_with_options(
            query_device_property_status_request, runtime)
        for T_T in return_value.body.data.list.property_status_info:
            # print(T_T)
            # print(T_T.name, T_T.value)
            DATA_D[T_T.name] = T_T.value
            if (T_T.name == '罐存厘米数'):
                DATA_L.append(DATA_D.copy())
                DATA_D.clear()
                DATA_D['时间'] = time_transform.timestamp_to_format(
                    int(T_T.time)/1000)
            DATA_L.append(DATA_D.copy())
            DATA_D.clear()

        return DATA_L


class History_data:

    def start_end_time(self) -> dict:
        timer = {}
        i = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print("当前的日期和时间是 %s" % i)
        now = time_transform.timeformat_to_timestamp(i)*1000
        print('现在的时间戳', now)
        Yesterday = time_transform.timeformat_to_timestamp(i)*1000-86400000
        print("昨天的时间戳", Yesterday)
        timer['昨天'] = Yesterday
        timer['现在'] = now
        return timer

    # def __init__(self, starttime: int, endtime: int) -> None:
    #     self.starttime = starttime
    #     self.endtime = endtime

    def history_out(self, key_s: str, name_s: str, s_time: int, e_time: int) -> list:
        mid_data_list = []
        client = Sample.create_client(
            'LTAI4FbnNz12HT85LUPbrFnU', 'UdPN3kNhB8aVJvrrbfIyng5sHupNpA')
        query_device_property_data_request = iot_20180120_models.QueryDevicePropertyDataRequest(
            product_key=key_s,
            device_name=name_s,
            start_time=s_time,
            end_time=e_time,
            page_size=50,
            identifier='Distance',
            asc=1
        )
        runtime = util_models.RuntimeOptions()
        history_value = client.query_device_property_data_with_options(
            query_device_property_data_request, runtime)
        history_value1 = history_value.to_map()

        print(history_value1['body']['Data'].keys())
        # history_value1.keys
        print('=========================')
        for ky in history_value1['body']['Data']['List']['PropertyInfo']:
            # print(ky)
            # print(time_transform.timestamp_to_format(int(ky['Time'])/1000))
            ky['Time'] = time_transform.timestamp_to_format(
                int(ky['Time'])/1000)
            mid_data_list.append(ky)
        print(mid_data_list)
        return mid_data_list[0]

    @staticmethod
    def dsplit_dict(dict):
        result = []
        for key, value in dict.items():
            sub_dict = {key: value}
            result.append(sub_dict)
        return result


device1 = Data_snapshot(ProductKey='a1bw1zXB8k4', DeviceName='Z065P1')  # 实例化
na_data = device1.data_out(device1.PRODUCT(), device1.DEVICE())
oil_num = Data_snapshot.DE_dict(device1.DEVICE())

print(na_data)

print("成功")

ti = History_data()

# print(ti.start_end_time())


her = ti.history_out(device1.PRODUCT(), device1.DEVICE(),
                     ti.start_end_time()['昨天'], ti.start_end_time()['现在'])
print('victoy')
print(her)
# print('分解')
print(History_data.dsplit_dict(her))
print(type(History_data.dsplit_dict(her)))
# na_data.append(her['Time'])
# na_data.append(her['Value'])
# na_data.append(her[1])
Montage = na_data+History_data.dsplit_dict(her)+oil_num
print(Montage)


title1 = datetime.datetime.now().strftime('%m%d%H%M%S')
str_time = "TL"+str(title1)  # 拼接新建表名
print('创建', str_time, '表格成功')
# shell = openpyxl.Workbook()
shell = openpyxl.load_workbook('实时数据.xlsx')







b1 = shell.create_sheet(str_time, 0)
b1.column_dimensions['B'].width = 20
b1.column_dimensions['I'].width = 12
b1.cell(row=1, column=1, value='井号')
b1.cell(row=1, column=2, value='Time')
b1.cell(row=1, column=3, value='Distance')
b1.cell(row=1, column=4, value='Distance1')
b1.cell(row=1, column=5, value='Distance2')
b1.cell(row=1, column=6, value='valu')
b1.cell(row=1, column=7, value='amp')
b1.cell(row=1, column=8, value='temp')
b1.cell(row=1, column=9, value='electricity')
# print('---------')
# mid_data = []
# for uu in her:
#     mid_data.append(uu)
# print('=========')
# for abc in mid_data:
#     print(abc)


# print("起始时间应早于终止时间,时间格式:xxxx-xx-xx xx:xx:xx")
# i = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# print("当前的日期和时间是 %s" % i)
# format = i
# # format = input("请输入起始时间：")
# # D = input("请输入要查询的天数：")
# # D = int(D)
# # # ad = time_transform.timeformat_to_timestamp(format)*1000
# # ad = time_transform.timeformat_to_timestamp(format)*1000 - 89600000*D
# ad = time_transform.timeformat_to_timestamp(format)*1000 - 89600000
# print(ad)
# # format = input("请输入终止时间：")

# ap = time_transform.timeformat_to_timestamp(format)*1000
# print(ap)
# # ac = ap-ad  # 计算查询时间差
# client = Sample.create_client(
#     'LTAI4FbnNz12HT85LUPbrFnU', 'UdPN3kNhB8aVJvrrbfIyng5sHupNpA')
# query_device_property_data_request = iot_20180120_models.QueryDevicePropertyDataRequest(
#     product_key='a1bw1zXB8k4',
#     device_name='Z065P6',
#     start_time=ad,
#     end_time=ap,
#     page_size=10,
#     identifier='Distance',
#     asc=1
# )
# runtime = util_models.RuntimeOptions()
# history_value = client.query_device_property_data_with_options(
#     query_device_property_data_request, runtime)

# for libra in history_value.body.data.list.property_info:
#     timer1 = time_transform.timestamp_to_format(libra.time/1000)
#     libra.time = timer1
#     # print('开始')
#     print(libra)
#     # for key in libra.items():
#     #     pass

# print(history_value.body.data.list.property_info)


# print(history_value.body.data.next_time)
# print(history_value.body.data.next_valid)

# {'headers': {'date': 'Wed, 10 May 2023 09:54:05 GMT', 'content-type': 'application/json;charset=utf-8', 'content-length': '534', 'connection': 'keep-alive', 'access-control-allow-origin': '*', 'access-control-expose-headers': '*', 'x-acs-request-id': '31C15353-CE2A-5CAC-8A1C-C652E9C18B26', 'x-acs-trace-id': 'd56fb5b54466a8259bb3c2a8ad56703d'}, 'statusCode': 200, 'body': {'Code': '', 'Data': {'List': {'PropertyInfo': [{'Time': 1683085281777, 'Value': '50.3'}, {
#     'Time': 1683085338781, 'Value': '50.4'}, {'Time': 1683085395767, 'Value': '50.5'}, {'Time': 1683085452707, 'Value': '50.7'}, {'Time': 1683085509708, 'Value': '50.7'}, {'Time': 1683088882819, 'Value': '14.5'}, {'Time': 1683088939802, 'Value': '11.3'}, {'Time': 1683088996802, 'Value': '10.5'}, {'Time': 1683089053745, 'Value': '10.3'}, {'Time': 1683089110734, 'Value': '10.3'}]}, 'NextTime': 1683089110736, 'NextValid': True}, 'RequestId': '31C15353-CE2A-5CAC-8A1C-C652E9C18B26', 'Success': True}}
