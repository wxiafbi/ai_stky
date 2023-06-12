import openpyxl as op


# [{'校验值': '221.6'}, {'传感器反射强度': '0'}, {'罐存厘米数1': '14.8'}, {'罐存厘米数2': '14.9'}, {'电压': '3.59'}, {'传 感器相对温度': '0.0'}, {'罐存厘米数': '14.8'}, {'时间': '2023-06-08 08:34:51'}, {'人工计量值': None}, {'Time': '2023-06-07 09:01:03'}, {'Value': '22.0'}, {'井号': 'Z065P1'}]
a1 = [{'a': 'acc1'}, {'b': 'acc1'}, {
    'c': 'acc1'}, {'d': 'acc1'}, {'e': 'acc1'}]
a2 = [{'a': 'acc2'}, {'b': 'acc2'}, {
    'c': 'aBC2'}, {'d': 'acc2'}, {'e': 'acc2'}]
a3 = [{'a': 'acc3'}, {'b': 'acc3'}, {
    'c': 'acc3'}, {'d': 'acc3'}, {'e': 'acc3'}]
a4 = [{'a': 'acc4'}, {'b': 'acc4'}, {
    'c': 'acc4'}, {'d': 'acc4'}, {'e': 'acc5'}]
key_1=['a','b','c','d','e']
tatal = []
tatal.append(a1)
tatal.append(a2)
tatal.append(a3)
tatal.append(a4)
print(tatal)
print(len(tatal[0]))
WORK = op.load_workbook('test.xlsx')
sheet = WORK.create_sheet('一区', 0)
# for i in 5:
#     print(5)
sheet.cell(row=1, column=1, value=key_1[0])
sheet.cell(row=1, column=2, value=key_1[1])
sheet.cell(row=1, column=3, value=key_1[2])
sheet.cell(row=1, column=4, value=key_1[3])
sheet.cell(row=1, column=5, value=key_1[4])
for i in range(len(tatal)):
    print('第一层',i)
    for j in range(len(a1)):
        print('第二层',j)
        sheet.cell(row=i+2, column=j+1, value=tatal[i][j][key_1[j]])
WORK.save('test.xlsx')
print(sheet)

